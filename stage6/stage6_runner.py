"""
stage6/stage6_runner.py
-----------------------
Stage 6 – Forced Alignment Scoring (WhisperX)
=============================================

Entry point consumed by main_pipeline.py.

Inputs
------
audio_path    : path to the cleaned WAV produced by stage 1
stage5_output : dict returned by run_stage5()
language      : ISO-639-1 language code from stage 2 (e.g. "ar", "en")
device        : "cuda" | "cpu"

Returns
-------
dict with the following keys:

Core metrics
  word_alignment_ratio      float [0,1]   – fraction of words with timestamps
  timing_deviation          dict          – mean/std/max/p90 (seconds)
  unaligned_segment_ratio   float [0,1]   – fraction of fully-unaligned segs
  avg_alignment_confidence  float [0,1]   – mean WhisperX word score
  phoneme_confidence        dict | None   – character-level score distribution

Anomaly detection
  hallucinated_segments     list[dict]    – segments flagged as hallucinations
  skipped_regions           list[dict]    – audio gaps with no word coverage
  overlapping_misalignments list[dict]    – word pairs with overlapping windows

Composite
  alignment_quality_score   float [0,1]   – single AQS summary
  per_segment_confidence    list[dict]    – per-segment confidence breakdown

Raw
  aligned_segments          list[dict]    – raw WhisperX segment output
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Optional

import librosa
import numpy as np

from .aligner      import WhisperXAligner
from .metrics      import (
    compute_word_alignment_ratio,
    compute_timing_deviation,
    compute_unaligned_segment_ratio,
    compute_avg_alignment_confidence,
    compute_phoneme_confidence,
    compute_per_segment_confidence,
)
from .hallucination import (
    detect_hallucinated_segments,
    detect_skipped_regions,
    detect_overlapping_misalignments,
)
from .pause_alignment import compute_punctuation_pause_score

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public API  ── single transcript
# ---------------------------------------------------------------------------

def run_stage6(
    audio_path:          str,
    stage5_output:       Optional[dict],
    language:            str,
    device:              str = "cuda",
    skip_gap_sec:        float = 2.0,
    provided_transcript: Optional[str] = None,
    _aligner:            Optional["WhisperXAligner"] = None,
) -> dict:
    """
    Forced-alignment scoring for a SINGLE transcript.

    Parameters
    ----------
    audio_path          : Cleaned WAV produced by Stage 1.
    stage5_output       : Output of run_stage5(). Ignored when
                          *provided_transcript* is set.
    language            : ISO-639-1 code (``"ar"``, ``"en"``, …).
    device              : ``"cuda"`` or ``"cpu"``.
    skip_gap_sec        : Min silence gap (s) to flag as a skipped region.
    provided_transcript : Align this transcript instead of Stage-5 output.
    _aligner            : Pre-loaded WhisperXAligner (avoids reloading the
                          alignment model when scoring many options).

    Returns
    -------
    dict – see module docstring.
    """

    # ── 0. Audio duration ─────────────────────────────────────────────────
    audio_dur = _audio_duration(audio_path)

    # ── 1. Build segment list ─────────────────────────────────────────────
    if provided_transcript is not None:
        # Score this specific transcript against the audio
        raw_segments: list[dict] = [{
            "start": 0.0,
            "end":   audio_dur,
            "text":  provided_transcript,
        }]
    else:
        details = stage5_output.get("details", []) if stage5_output else []
        if details and details[0].get("segments"):
            raw_segments = details[0]["segments"]
        else:
            # Fallback: whole-file segment from the consensus transcript
            reference = (stage5_output or {}).get("reference_transcript", "")
            raw_segments = [{"start": 0.0, "end": audio_dur, "text": reference}]
            logger.warning(
                "Stage 5 produced no word-level segments; "
                "using single full-span fallback (%.1f s).", audio_dur,
            )

    # ── 2. Forced alignment via WhisperX ──────────────────────────────────
    _lang = _safe_lang(language)
    aligner = _aligner or WhisperXAligner(language=_lang, device=device)

    logger.info("Stage 6 ▶ aligning %d segment(s) …", len(raw_segments))
    aligned_result   = aligner.align(audio_path, raw_segments)
    aligned_segments = aligned_result.get("segments", [])
    logger.info("Stage 6 ▶ alignment done (%d segments).", len(aligned_segments))

    # ── 3. Core metrics ─────────────────────────────────────────────────────
    word_ratio      = compute_word_alignment_ratio(aligned_segments)
    timing_dev      = compute_timing_deviation(raw_segments, aligned_segments)
    unaligned_ratio = compute_unaligned_segment_ratio(aligned_segments)
    avg_conf        = compute_avg_alignment_confidence(aligned_segments)
    phoneme_conf    = compute_phoneme_confidence(aligned_segments)
    per_seg_conf    = compute_per_segment_confidence(aligned_segments)

    logger.info(
        "Stage 6 metrics │ word_ratio=%.3f │ timing_dev_mean=%.3fs │ "
        "unaligned_ratio=%.3f │ avg_conf=%.3f",
        word_ratio, timing_dev["mean"], unaligned_ratio, avg_conf,
    )

    # ── 4. Anomaly detection ──────────────────────────────────────────────
    hallucinated = detect_hallucinated_segments(
        raw_segments, aligned_segments, audio_dur
    )
    skipped  = detect_skipped_regions(aligned_segments, audio_dur,
                                      min_gap=skip_gap_sec)
    overlaps = detect_overlapping_misalignments(aligned_segments)

    logger.info(
        "Stage 6 anomalies │ hallucinated_segs=%d │ skipped_regions=%d │ "
        "overlapping_words=%d",
        len(hallucinated), len(skipped), len(overlaps),
    )

    # ── 5a. Fine-grained discriminant signals ───────────────────────────────
    # Word-confidence std: spread of WhisperX per-word scores.
    # Two near-identical transcripts often differ here even when means tie.
    _word_scores = [
        w["score"]
        for seg in aligned_segments
        for w in seg.get("words", [])
        if w.get("score") is not None
    ]
    word_conf_std = float(np.std(_word_scores)) if len(_word_scores) > 1 else 0.0

    # Phoneme-level (character CTC) mean confidence – finer than word mean.
    phoneme_conf_mean = (
        phoneme_conf["mean"]
        if phoneme_conf and phoneme_conf.get("n_chars", 0) > 0
        else avg_conf
    )

    # ── 5b. Punctuation–pause alignment score ─────────────────────────────
    # Measures whether ؟ ، . markers coincide with real audio silences.
    punct_pause_result = compute_punctuation_pause_score(
        aligned_segments = aligned_segments,
        audio_path       = audio_path,
    )
    punct_pause_score = punct_pause_result["score"]
    logger.info(
        "Stage 6 punct-pause │ score=%.3f  gap_ratio=%.2f  coverage=%.2f  f1=%.3f",
        punct_pause_score,
        punct_pause_result["gap_ratio"],
        punct_pause_result["punct_coverage"],
        punct_pause_result["f1"],
    )

    # ── 5c. Composite Alignment Quality Score (AQS) ───────────────────────
    aqs = _compute_aqs(
        word_ratio        = word_ratio,
        avg_conf          = avg_conf,
        unaligned_ratio   = unaligned_ratio,
        timing_dev_mean   = timing_dev["mean"],
        timing_dev_p90    = timing_dev["p90"],
        phoneme_conf_mean = phoneme_conf_mean,
        word_conf_std     = word_conf_std,
        punct_pause_score = punct_pause_score,
        n_hallucinated    = len(hallucinated),
        skip_fraction     = sum(r["duration"] for r in skipped) / (audio_dur + 1e-8),
        n_overlaps        = len(overlaps),
    )
    logger.info("Stage 6 ▶ AQS = %.4f", aqs)

    # ── 6. Pack and return ────────────────────────────────────────────────
    return {
        # Core metrics
        "word_alignment_ratio":     word_ratio,
        "timing_deviation":         timing_dev,
        "unaligned_segment_ratio":  unaligned_ratio,
        "avg_alignment_confidence": avg_conf,
        "phoneme_confidence":       phoneme_conf,

        # Punctuation–pause alignment
        "punctuation_pause_score":  punct_pause_result,

        # Anomalies
        "hallucinated_segments":     hallucinated,
        "skipped_regions":           skipped,
        "overlapping_misalignments": overlaps,

        # Composite
        "alignment_quality_score": aqs,
        "per_segment_confidence":  per_seg_conf,

        # Raw output (for downstream stages or inspection)
        "aligned_segments": aligned_segments,
        "audio_duration":   audio_dur,
    }


# ---------------------------------------------------------------------------
# Public API  ── Excel multi-option scoring
# ---------------------------------------------------------------------------

def run_stage6_excel_options(
    audio_wav:    str,
    excel_path:   str,
    audio_id:     str | int,
    language:     str,
    device:       str = "cuda",
    skip_gap_sec: float = 2.0,
) -> dict:
    """
    Score all transcript options (option_1 … option_5) from *excel_path*
    against *audio_wav* using WhisperX forced alignment.

    The alignment model is loaded **once** and reused across all options so
    that GPU memory is loaded only a single time.

    Parameters
    ----------
    audio_wav   : Cleaned WAV produced by Stage 1 for this audio_id.
    excel_path  : Path to transcripts.xlsx (must have headers audio_id,
                  language, option_1 … option_5, [correct_option]).
    audio_id    : Row identifier matching the ``audio_id`` column.
    language    : ISO-639-1 code for the alignment model.
    device      : ``"cuda"`` or ``"cpu"``.
    skip_gap_sec: Min silence gap (s) to flag as a skipped region.

    Returns
    -------
    dict with keys:
        options          : {"option_1": {metrics…}, …}   per-option Stage-6 output
        ranked           : list of option keys sorted best→worst by AQS
        best_option      : str  e.g. ``"option_3"``
        best_aqs         : float
        correct_option   : int | None   from the Excel sheet (if present)
        language_tag     : str | None   raw language column value
        audio_id         : str
    """
    raw_options, correct_opt, lang_tag, cell_refs, excel_row = _load_excel_options(
        excel_path, str(audio_id)
    )

    _lang = _safe_lang(language)

    # ── Load alignment model ONCE for all options ──────────────────────────
    logger.info(
        "Stage 6 Excel ▶ loading alignment model  lang=%r  device=%s", _lang, device
    )
    aligner = WhisperXAligner(language=_lang, device=device)

    results: dict[str, Optional[dict]] = {}
    for key, text in raw_options.items():
        cell = cell_refs.get(key, "?")
        if not text:
            logger.info("Stage 6 Excel ▶ %s (cell %s) is empty — skipping.", key, cell)
            results[key] = None
            continue

        logger.info("Stage 6 Excel ▶ aligning %s  [Excel cell: %s] …", key, cell)
        try:
            out = run_stage6(
                audio_path          = audio_wav,
                stage5_output       = None,
                language            = _lang,
                device              = device,
                skip_gap_sec        = skip_gap_sec,
                provided_transcript = text,
                _aligner            = aligner,   # reuse loaded model
            )
            out["transcript"] = text
            out["excel_cell"] = cell          # e.g. "D2"
            out["excel_row"]  = excel_row     # e.g. 2
            results[key] = out
        except Exception as exc:  # noqa: BLE001
            logger.error("Stage 6 Excel ▶ %s (cell %s) failed: %s", key, cell, exc, exc_info=True)
            results[key] = None

    # ── Rank by AQS ───────────────────────────────────────────────────────
    scored = {
        k: v["alignment_quality_score"]
        for k, v in results.items()
        if v is not None
    }
    ranked = sorted(scored, key=scored.get, reverse=True)  # type: ignore[arg-type]
    best_key = ranked[0] if ranked else None

    return {
        "options":        results,
        "ranked":         ranked,
        "best_option":    best_key,
        "best_aqs":       scored.get(best_key) if best_key else None,
        "correct_option": correct_opt,
        "language_tag":   lang_tag,
        "audio_id":       str(audio_id),
        "cell_refs":      cell_refs,     # {"option_1": "D2", …}
        "excel_row":      excel_row,     # 2
        "excel_file":     excel_path,
    }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _audio_duration(audio_path: str) -> float:
    """Return audio duration in seconds using librosa (no FFmpeg required)."""
    y, sr = librosa.load(audio_path, sr=None, mono=True)
    return float(len(y) / sr)


def _safe_lang(language: str) -> str:
    """Return a valid ISO-639-1 code, falling back to English."""
    bad = {"unknown", "und", "xx", "", None}
    return language if language not in bad else "en"


def _load_excel_options(
    excel_path: str,
    audio_id:   str,
) -> tuple[dict[str, str], Optional[int], Optional[str], dict[str, str], int]:
    """
    Parse *excel_path* for the row matching *audio_id*.

    Returns
    -------
    options      : {"option_1": text, …, "option_5": text}  (empty string when cell is blank)
    correct_opt  : int | None  (from ``correct_option`` column if present)
    language_tag : str | None  (raw ``language`` column value)
    cell_refs    : {"option_1": "D2", …}  Excel cell address for each option
    excel_row    : int   1-based spreadsheet row where the audio_id was found
    """
    try:
        import openpyxl  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read the Excel file.\n"
            "Install with:  pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None)
    if raw_headers is None:
        raise ValueError(f"{excel_path}: workbook is empty.")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    idx: dict[str, int] = {h: i for i, h in enumerate(headers) if h}

    # Validate required columns
    required = {"audio_id", "option_1", "option_2", "option_3", "option_4", "option_5"}
    missing = required - idx.keys()
    if missing:
        raise ValueError(
            f"{excel_path}: missing required column(s): {', '.join(sorted(missing))}\n"
            f"Found headers: {headers}"
        )

    options: dict[str, str] = {}
    cell_refs: dict[str, str] = {}
    correct_opt: Optional[int] = None
    language_tag: Optional[str] = None
    found = False
    excel_row = -1

    # iter_rows starts at row 2 (row 1 was the header consumed above)
    for data_row_idx, row in enumerate(rows, start=2):
        cell_id = row[idx["audio_id"]]
        if cell_id is None:
            continue
        if str(cell_id).strip() == audio_id:
            found = True
            excel_row = data_row_idx
            # Language tag
            if "language" in idx:
                language_tag = str(row[idx["language"]]).strip() if row[idx["language"]] is not None else None
            # option_1 … option_5  + build cell references like "D2"
            for k in range(1, 6):
                col = f"option_{k}"
                col_num = idx[col]           # 0-based column index
                col_letter = get_column_letter(col_num + 1)   # 1-based for openpyxl
                cell_refs[col] = f"{col_letter}{excel_row}"
                val = row[col_num] if idx.get(col) is not None else None
                options[col] = str(val).strip() if val is not None else ""
            # correct_option (optional column)
            if "correct_option" in idx:
                raw_correct = row[idx["correct_option"]]
                try:
                    correct_opt = int(raw_correct) if raw_correct is not None else None
                except (TypeError, ValueError):
                    correct_opt = None
            break

    wb.close()

    if not found:
        raise ValueError(f"{excel_path}: audio_id={audio_id!r} not found.")

    return options, correct_opt, language_tag, cell_refs, excel_row


def _compute_aqs(
    *,
    word_ratio:        float,
    avg_conf:          float,
    unaligned_ratio:   float,
    timing_dev_mean:   float,
    timing_dev_p90:    float,
    phoneme_conf_mean: float,
    word_conf_std:     float,
    punct_pause_score: float,
    n_hallucinated:    int,
    skip_fraction:     float,
    n_overlaps:        int,
) -> float:
    """
    Alignment Quality Score (AQS) – a single [0, 1] discriminating summary
    designed to separate nearly-identical transcripts that differ in
    punctuation placement and word boundaries.

    Component weights (sum = 1.0)
    -----------------
    0.22 × word_alignment_ratio        – fraction of words with timestamps
    0.15 × avg_alignment_confidence    – mean WhisperX word-level score
    0.10 × (1 − unaligned_seg_ratio)   – segment-level coverage
    0.09 × timing_score_mean           – 1 − mean_dev/3 (Whisper boundary fit)
    0.06 × timing_score_p90            – 1 − p90_dev/3  (tail alignment fit)
    0.13 × phoneme_conf_score          – char-level CTC confidence (finer)
    0.05 × consistency_score           – 1 − norm(word_conf_std)
    0.20 × punct_pause_score           – punctuation ↔ audio-silence alignment

    Penalties (capped)
    ------------------
    −0.05 per hallucinated segment  (max −0.30)
    −0.50 × skip_fraction           (proportional to skipped audio)
    −0.01 per overlapping word pair (max −0.10)
    """
    # Timing scores: 1.0 at 0 s deviation, 0.0 at ≥ 3 s deviation
    timing_score_mean = float(np.clip(1.0 - timing_dev_mean / 3.0, 0.0, 1.0))
    timing_score_p90  = float(np.clip(1.0 - timing_dev_p90  / 3.0, 0.0, 1.0))

    # Phoneme score is already in [0, 1]
    phoneme_score = float(np.clip(phoneme_conf_mean, 0.0, 1.0))

    # Consistency: low word-conf std → uniform alignment → 1.0.
    # Std of 0.25 maps to 0 (fully inconsistent for typical CTC outputs).
    consistency_score = float(np.clip(1.0 - word_conf_std / 0.25, 0.0, 1.0))

    base = (
        0.22 * word_ratio
        + 0.15 * avg_conf
        + 0.10 * (1.0 - unaligned_ratio)
        + 0.09 * timing_score_mean
        + 0.06 * timing_score_p90
        + 0.13 * phoneme_score
        + 0.05 * consistency_score
        + 0.20 * float(np.clip(punct_pause_score, 0.0, 1.0))
    )

    hallucination_penalty = min(0.30, n_hallucinated * 0.05)
    skip_penalty          = min(0.30, skip_fraction * 0.50)
    overlap_penalty       = min(0.10, n_overlaps * 0.01)

    aqs = base - hallucination_penalty - skip_penalty - overlap_penalty
    return float(np.clip(aqs, 0.0, 1.0))
