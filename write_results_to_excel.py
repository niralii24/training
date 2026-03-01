"""
write_results_to_excel.py
-------------------------
Runs the full pipeline (stages 1-2, 5-8) for every audio file found in
downloaded_audio/ and writes per-option metrics plus a 'detected_option'
column back into transcripts.xlsx.

New columns appended (right of existing data)
--------------------------------------------
For each option N in 1..5:
  option_N_aqs    – Alignment Quality Score  (Stage 6)
  option_N_wer    – Mean Word Error Rate      (Stage 7)
  option_N_cer    – Mean Char Error Rate      (Stage 7)
  option_N_tss    – Transcript Similarity     (Stage 7)
  option_N_lgs    – Linguistic Grammar Score  (Stage 8)
  option_N_final  – Weighted final score

Then:
  detected_option – 1-based int of the pipeline's top-ranked option
"""
import sys, os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ["PYTHONUTF8"] = "1"

import yaml
import torchaudio
import openpyxl
from openpyxl.styles import PatternFill, Font

from stage1.stage1_runner     import run_stage1
from stage2.language_detector import detect_language
from stage5.stage5_runner     import run_stage5
from stage6.stage6_runner     import run_stage6_excel_options
from stage7.stage7_runner     import run_stage7
from stage8.stage8_runner     import run_stage8, compute_final_score


# ── helpers ────────────────────────────────────────────────────────────────

def load_config(path="config.yaml"):
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg


def resolve_device(d, default="cuda"):
    if d == "auto":
        import torch
        return "cuda" if torch.cuda.is_available() else "cpu"
    return d


# ── per-row pipeline ────────────────────────────────────────────────────────

def run_pipeline_for_row(audio_id, cfg):
    """
    Run stages 1-2, 5-8 for one audio file.

    Returns
    -------
    dict keyed by option name ('option_1' … 'option_5') with sub-dict:
        aqs, wer, cer, tss, lgs, final_score
    and top-level key 'detected_option' (1-based int).
    Returns None if audio file not found.
    """
    audio_dir = cfg["paths"]["audio_dir"]
    clean_wav = cfg["paths"]["clean_wav"]
    excel_file = cfg["paths"]["excel_file"]
    weights   = cfg["final_score"]["weights"]

    audio_file = os.path.join(audio_dir, f"{audio_id}.mp3")
    if not os.path.exists(audio_file):
        print(f"  [SKIP] {audio_file} not found.")
        return None

    # Stage 1
    print(f"  [1] Loading audio …")
    s1 = run_stage1(audio_file)
    waveform, sr = s1["waveform"], s1["sample_rate"]
    torchaudio.save(clean_wav, waveform, sr, encoding="PCM_S", bits_per_sample=16)

    # Stage 2
    print(f"  [2] Detecting language …")
    lang, conf, _, method = detect_language(waveform, sr)
    print(f"      lang={lang}  conf={conf:.2%}  method={method}")

    # Stage 5
    print(f"  [5] ASR transcription …")
    s5 = run_stage5(clean_wav, cfg["stage5"], language=lang)

    # Stage 6
    print(f"  [6] Forced alignment scoring …")
    s6_device   = resolve_device(cfg["stage6"]["device"])
    s6_skip_gap = cfg["stage6"]["skip_gap_sec"]
    s6 = run_stage6_excel_options(
        audio_wav    = clean_wav,
        excel_path   = excel_file,
        audio_id     = audio_id,
        language     = lang,
        device       = s6_device,
        skip_gap_sec = s6_skip_gap,
    )
    correct_opt = s6["correct_option"]

    # Stage 7
    print(f"  [7] Transcript similarity …")
    excel_opts = {
        k: (v["transcript"] if v is not None else "")
        for k, v in s6["options"].items()
    }
    s7 = run_stage7(
        excel_options  = excel_opts,
        asr_references = s5["reference_transcripts"],
        language       = lang,
        correct_option = correct_opt,
    )

    # Stage 8
    print(f"  [8] Linguistic grammar scoring …")
    s8_device = resolve_device(cfg["stage8"]["device"])
    s8 = run_stage8(
        excel_options  = excel_opts,
        language       = lang,
        device         = s8_device,
        correct_option = correct_opt,
    )

    # Collate per-option results
    option_results = {}
    final_scores   = {}

    for i in range(1, 6):
        key = f"option_{i}"
        s6r = s6["options"].get(key)
        s7r = s7["options"].get(key)
        s8r = s8["options"].get(key)

        aqs   = s6r["alignment_quality_score"] if s6r else None
        wer_v = s7r["mean_wer"]               if s7r else None
        cer_v = s7r["mean_cer"]               if s7r else None
        tss   = s7r["tss"]                    if s7r else None
        lgs   = s8r["lgs"]                    if s8r else None

        fin = None
        if aqs is not None and tss is not None and lgs is not None:
            fin = compute_final_score(aqs, tss, lgs, weights)
            final_scores[key] = fin

        option_results[key] = {
            "aqs":         aqs,
            "wer":         wer_v,
            "cer":         cer_v,
            "tss":         tss,
            "lgs":         lgs,
            "final_score": fin,
        }

    detected = None
    if final_scores:
        best_key = max(final_scores, key=final_scores.get)
        detected = int(best_key.split("_")[1])

    option_results["detected_option"] = detected
    return option_results


# ── Excel writer ─────────────────────────────────────────────────────────────

# Metrics in the order they appear per option
METRICS = ["wer", "aqs", "cer", "tss", "lgs", "final_score"]
OPTIONS = [f"option_{i}" for i in range(1, 6)]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
DETECTED_FILL = PatternFill("solid", fgColor="375623")
HEADER_FONT   = Font(bold=True, color="FFFFFF")


def add_or_update_columns(xlsx_path, all_results: dict):
    """
    Open xlsx_path, add new metric columns (or overwrite if already present),
    and save.

    all_results : { audio_id (str) -> {option_N: {metric: value}, detected_option: int} }
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Build header map  { col_name_lower -> col_index (1-based) }
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(h).lower().strip(): i+1 for i, h in enumerate(header_row) if h is not None}

    # Find audio_id column
    aid_col = header_map.get("audio_id", 1)

    # Determine new column headers we need
    # Grouped by metric: option_1_wer, option_2_wer … option_5_wer | option_1_cer … | …
    new_headers = []
    for m in METRICS:
        for opt in OPTIONS:
            new_headers.append(f"{opt}_{m}")
    new_headers.append("detected_option")

    # Map each new header to its column index (extend if necessary)
    next_free = ws.max_column + 1
    col_index = {}
    for h in new_headers:
        h_lower = h.lower()
        if h_lower in header_map:
            col_index[h] = header_map[h_lower]
        else:
            col_index[h] = next_free
            next_free += 1

    # Write headers
    for h, ci in col_index.items():
        cell = ws.cell(1, ci, value=h)
        if h == "detected_option":
            cell.fill = DETECTED_FILL
        else:
            cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Write data rows
    for row in range(2, ws.max_row + 1):
        audio_id_val = ws.cell(row, aid_col).value
        if audio_id_val is None:
            continue
        aid_str = str(int(audio_id_val)) if isinstance(audio_id_val, float) else str(audio_id_val).strip()

        result = all_results.get(aid_str)
        if result is None:
            continue

        for m in METRICS:
            for opt in OPTIONS:
                col_h = f"{opt}_{m}"
                val   = result.get(opt, {}).get(m)
                ws.cell(row, col_index[col_h], value=round(val, 5) if val is not None else None)

        ws.cell(row, col_index["detected_option"], value=result.get("detected_option"))

    wb.save(xlsx_path)
    print(f"\n  Saved results -> {xlsx_path}")


# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    cfg = load_config("config.yaml")

    # Set HF cache env vars
    cache_dir = cfg["global"]["cache_dir"]
    os.environ["HF_HOME"]               = cache_dir
    os.environ["HUGGINGFACE_HUB_CACHE"] = cache_dir

    excel_path = cfg["paths"]["excel_file"]
    audio_dir  = cfg["paths"]["audio_dir"]

    # Discover audio IDs from downloaded_audio/ (1.mp3, 2.mp3, …)
    import re
    audio_ids = sorted(
        int(m.group(1))
        for f in os.listdir(audio_dir)
        if (m := re.fullmatch(r"(\d+)\.mp3", f))
    )
    print(f"Found audio files: {audio_ids}")

    all_results = {}
    for audio_id in audio_ids:
        print(f"\n{'='*60}")
        print(f"  audio_id = {audio_id}")
        print(f"{'='*60}")
        try:
            res = run_pipeline_for_row(audio_id, cfg)
            if res is not None:
                all_results[str(audio_id)] = res
                detected = res.get("detected_option")
                print(f"  -> detected_option = {detected}")
                # ── Write this row to Excel immediately ──────────────
                print(f"  Writing audio_id={audio_id} to {excel_path} …")
                add_or_update_columns(excel_path, {str(audio_id): res})
        except Exception as exc:
            import traceback
            print(f"  [ERROR] audio_id={audio_id}: {exc}")
            traceback.print_exc()

    if not all_results:
        print("No results to write.")
